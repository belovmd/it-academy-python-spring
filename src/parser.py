from bs4 import BeautifulSoup as bsoup
import logging
from openpyxl import Workbook
import requests



def html_request(html_url_, headers):
    """Request function

    This is html_request function connects to sever with logging parameters
    of successful or unsuccessful connection
    """
    logging.basicConfig(
        filename="urlopen.log",
        level=logging.INFO,
        format='%(asctime)s-%(filename)s-%(funcName)s-'
               '%(name)s-%(levelname)s-%(message)s',
        filemode='w'
    )

    session = requests.Session()

    try:
        request = session.get(html_url_, headers=headers, timeout=5)
        logging.info(
            f'Successfully connected to url Server code answer '
            f'is {request.status_code}'
        )
        return request
    except requests.ConnectionError as e:
        logging.error(f'Connection ERROR - {e}')
    except requests.Timeout as e:
        logging.error(f'Timeout ERROR - {e}')
    except requests.RequestException as e:
        logging.error(f'General ERROR - {e}')
    return


def parse_html(html_url_, headers):
    """Parser of 'Jobs.tut.by'

    Parse_html function collects all vacations from jobs.tut.by
    from given number of pages (maximum_number_of _pages parameter)
    with:
        title,
        link of vacancy,
        salary,
        company_name,
        company_address,
        full description from vacancy page
    :param html_url_: Url for parsing
    :param headers: Headers to imitate browser
    :return: list of dictionaries with all data from list above
    """

    page_number = 0
    maximum_number_of_pages = 2
    all_data = []
    vacancy_pr_divs_data_attrs = {
        'data-qa': 'vacancy-serp__vacancy vacancy-serp__vacancy_premium'
    }
    vacancy_divs_data_attrs = {'data-qa': 'vacancy-serp__vacancy'}
    title_div_attrs = {'class': 'bloko-link HH-LinkModifier'}
    description_href_div_attrs = {'class': 'bloko-link HH-LinkModifier'}
    salary_div_attrs = {'class': 'vacancy-serp-item__sidebar'}
    company_name_div_attrs = {'class': 'vacancy-serp-item__meta-info'}
    address_div_attrs = {'data-qa': 'vacancy-serp__vacancy-address'}
    description_div_attrs = {
        'class': 'g-user-content', 'data-qa': 'vacancy-description'
    }

    #  We can use "True" condition to parse all pages
    #  For this reason I used "while"
    while page_number < maximum_number_of_pages:

        html_content = html_request(html_url_ + str(page_number), headers)

        if not html_content:
            return all_data

        soup_root = bsoup(html_content.content, 'html.parser')
        divs_pr = soup_root.find_all('div', attrs=vacancy_pr_divs_data_attrs)
        divs_s = soup_root.find_all('div', attrs=vacancy_divs_data_attrs)
        divs = divs_pr + divs_s

        for div_ind, div in enumerate(divs):

            title = div.find('a', attrs=title_div_attrs).text
            href = div.find('a', attrs=description_href_div_attrs)['href']
            salary = div.find('div', attrs=salary_div_attrs).text
            company = div.find('div', attrs=company_name_div_attrs).text
            address = div.find('span', attrs=address_div_attrs).text

            resp_and_req = html_request(href, headers)
            all_text = []

            if resp_and_req:
                soup_description = bsoup(resp_and_req.content, 'html.parser')
                div_vacancy_description = soup_description.find_all(
                    'div', attrs=description_div_attrs
                )
                if div_vacancy_description:

                    for row in div_vacancy_description[0].find_all('li'):
                        all_text.append(row.text)

            all_data.append(
                {
                    'title': title,
                    'salary': salary,
                    'company': company,
                    'description': all_text,
                    'address': address,
                    'href': href
                }
            )

        page_number += 1
    return all_data


def xls_save(file_name_, html_url_, headers):
    """Data file creator

    Create ".xlsx" file with all data from parsing page:
        Index of vacancy
        Title of vacancy
        Company name
        Salary
        Link of vacancy
        Full description of vacancy

    :param file_name_: Name of ".xlsx" file for parsing data
    :param html_url_: Url for parsing
    :param headers: Headers to imitate browser
    :return: None
    """

    all_data = parse_html(html_url_, headers)
    try:
        jobs_data = Workbook()
        sheet = jobs_data.active
        sheet.title = 'ALL_DATA'

        sheet.cell(row=1, column=1).value = 'Id'
        sheet.cell(row=1, column=2).value = 'Title'
        sheet.cell(row=1, column=3).value = 'Company'
        sheet.cell(row=1, column=4).value = 'Salary'
        sheet.cell(row=1, column=5).value = 'Link to vacancy'
        sheet.cell(row=1, column=6).value = 'Description'

        for index, vacancy_data in enumerate(all_data):

            sheet.cell(row=index + 2, column=1).value = index + 1
            sheet.cell(row=index + 2, column=2).value = vacancy_data['title']
            sheet.cell(row=index + 2, column=3).value = vacancy_data['company']
            sheet.cell(row=index + 2, column=4).value = vacancy_data['salary']
            sheet.cell(row=index + 2, column=5).value = vacancy_data['href']
            sheet.cell(row=index + 2, column=6).value = (
                '\n'.join(vacancy_data['description'])
            )

        jobs_data.save(file_name_)

    except Exception:
        print(f'Error with open/save "{file_name_}" ')


if __name__ == '__main__':
    html_url = (
        'https://grodno.jobs.tut.by/search/vacancy?L_is_autosearch='
        'false&area=1002&clusters=true&currency_code='
        'BYR&enable_snippets=true&search_period=7&page='
    )
    headers = {
        'accept': (
            'text/html,application/xhtml+xml,application/xml;q='
            '0.9,image/webp,image/apng,*/*;q='
            '0.8,application/signed-exchange;v=b3;q=0.9'
        ),
        'user-agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/79.0.3945.88 Safari/537.36'
        )
    }
    xls_save('Jobs_tut.xlsx', html_url, headers)
